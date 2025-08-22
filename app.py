from flask import Flask, render_template, request, send_file, session, redirect, url_for
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = "your_secret_key_here"  # replace with a strong secret

EXCEL_FILE = 'charges_data.xlsx'

# Hardcoded username and password
USERNAME = "admin"
PASSWORD = "1234"

# Create Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Serial Number", "Vehicle Number", "Charges"])
    wb.save(EXCEL_FILE)

# -------------------- LOGIN ROUTE --------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == USERNAME and password == PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            return "❌ Invalid credentials. <a href='/login'>Try again</a>"
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

# -------------------- PROTECTED ROUTES --------------------
def login_required(func):
    def wrapper(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
@login_required
def submit():
    date = request.form['date']
    serial = request.form['serial']
    vehicle = request.form['vehicle']
    charges = float(request.form['charges'])

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([date, serial, vehicle, charges])
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        return "❌ Please close the Excel file before saving. <a href='/'>Go Back</a>"

    return "✅ Data saved successfully. <a href='/'>Go Back</a>"

@app.route('/download')
@login_required
def download():
    return send_file(EXCEL_FILE, as_attachment=True)

@app.route('/daily_total')
@login_required
def daily_total():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    records = data[1:]
    today = datetime.now().strftime('%Y-%m-%d')
    total = sum(row[3] for row in records if row[0] == today and isinstance(row[3], (int, float)))
    ws.append([f'Total for {today}', '', '', total])
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        return "❌ Please close the Excel file before saving. <a href='/'>Go Back</a>"

    return f"🧮 Total ₹{total} for {today} added. <a href='/'>Go Back</a>"

@app.route('/reset')
@login_required
def reset_excel():
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Serial Number", "Vehicle Number", "Charges"])
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        return "❌ Please close the Excel file before resetting. <a href='/'>Go Back</a>"

    return "🗑️ Excel file cleared. <a href='/'>Go Back</a>"

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
